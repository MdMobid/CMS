import sys, csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR = "lab_data"

HDR_FILL  = PatternFill("solid", start_color="1E293B")   # dark navy
HDR_FONT  = Font(bold=True, color="F59E0B", name="Calibri", size=11)
ALT_FILL  = PatternFill("solid", start_color="F1F5F9")
STD_FONT  = Font(name="Calibri", size=10)
TITLE_FONT= Font(bold=True, name="Calibri", size=14, color="1E293B")
CENTER    = Alignment(horizontal="center", vertical="center")
thin      = Side(style="thin", color="CBD5E1")
BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = CENTER; cell.border = BORDER

def style_data(ws, row, cols, alt=False):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        if alt: cell.fill = ALT_FILL
        cell.font = STD_FONT; cell.alignment = CENTER; cell.border = BORDER

def read_students(class_id):
    path = os.path.join(DATA_DIR, f"{class_id}_students.csv")
    students = {}
    if os.path.exists(path):
        with open(path) as f:
            for line in f:
                parts = line.strip().split(",", 1)
                if len(parts) == 2:
                    students[int(parts[0])] = parts[1]
    return students

def export_attendance(class_id, outfile):
    students = read_students(class_id)
    path = os.path.join(DATA_DIR, f"{class_id}_attendance.csv")

    # Parse attendance: {roll: {date: present}}
    data = {}
    dates = set()
    if os.path.exists(path):
        with open(path) as f:
            for line in f:
                parts = line.strip().split(",")
                if len(parts) == 3:
                    roll, date, present = int(parts[0]), parts[1], int(parts[2])
                    if roll not in data: data[roll] = {}
                    data[roll][date] = present
                    dates.add(date)

    dates = sorted(dates)
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Title
    ws.merge_cells(f"A1:{get_column_letter(3+len(dates))}1")
    ws["A1"] = f"Attendance Report — Class {class_id}"
    ws["A1"].font = TITLE_FONT; ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    # Header row
    headers = ["Roll", "Name"] + dates + ["Total Present", "Total Classes", "% Attendance"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=2, column=c, value=h)
    style_header(ws, 2, len(headers))
    ws.row_dimensions[2].height = 22

    # Data rows
    all_rolls = sorted(set(list(students.keys()) + list(data.keys())))
    for i, roll in enumerate(all_rolls):
        r = i + 3
        alt = (i % 2 == 1)
        ws.cell(row=r, column=1, value=roll)
        ws.cell(row=r, column=2, value=students.get(roll, f"Student {roll}"))
        for j, date in enumerate(dates):
            val = data.get(roll, {}).get(date, "")
            display = "P" if val == 1 else ("A" if val == 0 else "")
            cell = ws.cell(row=r, column=3+j, value=display)
            if display == "P":
                cell.fill = PatternFill("solid", start_color="D1FAE5")
                cell.font = Font(color="065F46", bold=True, name="Calibri", size=10)
            elif display == "A":
                cell.fill = PatternFill("solid", start_color="FEE2E2")
                cell.font = Font(color="991B1B", bold=True, name="Calibri", size=10)
            cell.alignment = CENTER; cell.border = BORDER
        # Totals
        date_cols = [3+j for j in range(len(dates))]
        tp_col = get_column_letter(3+len(dates))
        tc_col = get_column_letter(4+len(dates))
        pct_col= get_column_letter(5+len(dates))
        present = sum(1 for d in dates if data.get(roll, {}).get(d, -1) == 1)
        total   = len([d for d in dates if d in data.get(roll, {})])
        ws.cell(row=r, column=3+len(dates),   value=present)
        ws.cell(row=r, column=4+len(dates),   value=total)
        ws.cell(row=r, column=5+len(dates),   value=f"{(present/total*100):.1f}%" if total else "N/A")
        style_data(ws, r, len(headers), alt)
        # Re-apply P/A styling
        for j, date in enumerate(dates):
            val = data.get(roll, {}).get(date, "")
            if val in (0, 1):
                c = ws.cell(row=r, column=3+j)
                c.alignment = CENTER; c.border = BORDER

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 22
    for j in range(len(dates)):
        ws.column_dimensions[get_column_letter(3+j)].width = 12
    ws.column_dimensions[get_column_letter(3+len(dates))].width = 14
    ws.column_dimensions[get_column_letter(4+len(dates))].width = 14
    ws.column_dimensions[get_column_letter(5+len(dates))].width = 14

    ws.freeze_panes = "C3"
    wb.save(outfile)
    print(f"OK:{outfile}")

def export_marks(class_id, outfile):
    students = read_students(class_id)
    path = os.path.join(DATA_DIR, f"{class_id}_marks.csv")

    # {roll: {quiz_no: (marks, max_marks)}}
    data = {}; quizzes = {}
    if os.path.exists(path):
        with open(path) as f:
            for line in f:
                parts = line.strip().split(",")
                if len(parts) == 4:
                    roll, qno, marks, mx = int(parts[0]), int(parts[1]), float(parts[2]), float(parts[3])
                    if roll not in data: data[roll] = {}
                    data[roll][qno] = (marks, mx)
                    quizzes[qno] = mx

    quiz_nos = sorted(quizzes.keys())
    wb = Workbook()
    ws = wb.active
    ws.title = "Quiz Marks"

    total_cols = 2 + len(quiz_nos)*2 + 3
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    ws["A1"] = f"Quiz Marks Report — Class {class_id}"
    ws["A1"].font = TITLE_FONT; ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    headers = ["Roll", "Name"]
    for q in quiz_nos:
        headers += [f"Quiz {q} Marks", f"Quiz {q} Max"]
    headers += ["Total Marks", "Total Max", "Overall %"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=2, column=c, value=h)
    style_header(ws, 2, len(headers))
    ws.row_dimensions[2].height = 22

    all_rolls = sorted(set(list(students.keys()) + list(data.keys())))
    for i, roll in enumerate(all_rolls):
        r = i + 3
        alt = (i % 2 == 1)
        ws.cell(row=r, column=1, value=roll)
        ws.cell(row=r, column=2, value=students.get(roll, f"Student {roll}"))
        tot_m = 0; tot_mx = 0
        for j, qno in enumerate(quiz_nos):
            m, mx = data.get(roll, {}).get(qno, ("", quizzes[qno]))
            ws.cell(row=r, column=3+j*2,   value=m if m != "" else "—")
            ws.cell(row=r, column=3+j*2+1, value=mx)
            if m != "": tot_m += m; tot_mx += mx
        nc = 3 + len(quiz_nos)*2
        ws.cell(row=r, column=nc,   value=tot_m)
        ws.cell(row=r, column=nc+1, value=tot_mx if tot_mx else "—")
        pct = f"{tot_m/tot_mx*100:.1f}%" if tot_mx else "—"
        ws.cell(row=r, column=nc+2, value=pct)
        style_data(ws, r, len(headers), alt)

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 22
    for j in range(len(quiz_nos)*2 + 3):
        ws.column_dimensions[get_column_letter(3+j)].width = 14
    ws.freeze_panes = "C3"
    wb.save(outfile)
    print(f"OK:{outfile}")

if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("Usage: export_xlsx.py <attendance|marks> <class_id> <output.xlsx>")
        sys.exit(1)
    mode, class_id, outfile = sys.argv[1], sys.argv[2], sys.argv[3]
    if mode == "attendance": export_attendance(class_id, outfile)
    elif mode == "marks":    export_marks(class_id, outfile)
    else: print("ERROR:Unknown mode")
